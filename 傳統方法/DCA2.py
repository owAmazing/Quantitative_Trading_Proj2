import pandas as pd
import numpy as np
import numpy_financial as npf
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ====================== 設定 ======================
DESKTOP_PATH = os.path.expanduser("~/Desktop")
OUTPUT_DIR   = os.path.join(DESKTOP_PATH, "dcaproject")   # 輸出資料夾
TICKERS = ['2330.TW', '2454.TW', 'AAPL', 'PEP', 'KO']
RISK_FREE_RATE = 0.017  # 無風險利率（年化，定存）
TEST_START_YEARS = list(range(2007, 2027))

# ====================== 讀取 CSV ======================
def load_csv(ticker):
    filename = f"{ticker.replace('.','_')}_20y.csv"
    file_path = os.path.join(DESKTOP_PATH, filename)
    if not os.path.exists(file_path):
        print(f"✗ 找不到檔案：{filename}，請先執行下載程式。")
        return None
    df = pd.read_csv(file_path, parse_dates=['date'], index_col='date')
    df.dropna(subset=['close'], inplace=True)
    return df

# ====================== DCA 核心計算 ======================
def compute_dca(df):
    df = df.copy()
    df.index = pd.to_datetime(df.index)
    df['year_month'] = df.index.to_period('M')
    # 修正 FutureWarning：加上 include_groups=False
    first_trading_days = df.groupby('year_month').apply(
        lambda x: x.index.min(), include_groups=False
    )

    records = []
    cumulative_shares   = 0
    cumulative_invested = 0.0
    peak_value_1        = 0.0
    max_drawdown_1      = 0.0

    for month_num, (period, trade_date) in enumerate(first_trading_days.items(), start=1):
        close_price = df.loc[trade_date, 'close']
        cumulative_shares   += 1
        cumulative_invested += close_price

        portfolio_value = round(close_price * cumulative_shares, 2)

        if portfolio_value > peak_value_1:
            peak_value_1 = portfolio_value
        drawdown_1 = (portfolio_value - peak_value_1) / peak_value_1
        if drawdown_1 < max_drawdown_1:
            max_drawdown_1 = drawdown_1

        records.append({
            '月份':             month_num,
            '日期':             trade_date.strftime('%Y/%m/%d'),
            '當月股價':         round(close_price, 2),
            '投入金額':         round(close_price, 2),
            '買入股數':         1,
            '累積股數':         cumulative_shares,
            '目前累積投入金額': round(cumulative_invested, 2),
            '目前持有價值':     portfolio_value,
            '當期回撤':         drawdown_1,
        })

    result_df = pd.DataFrame(records)

    final_value  = result_df['目前持有價值'].iloc[-1]
    cash_flows   = [-row['投入金額'] for _, row in result_df.iterrows()]
    cash_flows[-1] += final_value
    monthly_irr  = npf.irr(cash_flows)
    annual_irr   = (1 + monthly_irr) ** 12 - 1

    total_invested    = result_df['投入金額'].sum()
    cumulative_return = (final_value - total_invested) / total_invested

    prices          = [r['當月股價'] for r in records]
    monthly_returns = [(prices[i] - prices[i-1]) / prices[i-1] for i in range(1, len(prices))]
    monthly_rf      = (1 + RISK_FREE_RATE) ** (1 / 12) - 1
    downside_diffs  = [min(r - monthly_rf, 0) for r in monthly_returns]
    downside_variance  = np.mean([d ** 2 for d in downside_diffs])
    downside_deviation = np.sqrt(downside_variance) * np.sqrt(12)
    sortino_ratio = (annual_irr - RISK_FREE_RATE) / downside_deviation if downside_deviation != 0 else float('nan')

    return (result_df, annual_irr, cumulative_return, total_invested,
            final_value, max_drawdown_1, sortino_ratio)

# ====================== 輸出逐月明細 Excel ======================
def save_excel(result_df, ticker, annual_irr, cumulative_return, total_invested,
               final_value, max_drawdown_1, sortino_ratio):
    wb = Workbook()
    ws = wb.active
    ws.title = "DCA績效"

    header_font  = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    header_fill  = PatternFill('solid', start_color='1F4E79')
    data_font    = Font(name='Arial', size=10)
    alt_fill     = PatternFill('solid', start_color='D6E4F0')
    summary_font = Font(name='Arial', bold=True, size=10, color='1F4E79')
    summary_fill = PatternFill('solid', start_color='EBF3FB')
    center       = Alignment(horizontal='center', vertical='center')
    right        = Alignment(horizontal='right',  vertical='center')
    thin_border  = Border(
        left=Side(style='thin',   color='B0C4DE'),
        right=Side(style='thin',  color='B0C4DE'),
        top=Side(style='thin',    color='B0C4DE'),
        bottom=Side(style='thin', color='B0C4DE'),
    )

    headers    = ['月份', '日期', '當月股價', '投入金額', '買入股數', '累積股數',
                  '目前累積投入金額', '目前持有價值', '當期回撤']
    col_widths = [8, 14, 14, 14, 10, 10, 18, 16, 12]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = center; cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[1].height = 22

    for row_idx, row in result_df.iterrows():
        excel_row = row_idx + 2
        fill = alt_fill if row_idx % 2 == 0 else None
        values = [row['月份'], row['日期'], row['當月股價'], row['投入金額'],
                  row['買入股數'], row['累積股數'], row['目前累積投入金額'],
                  row['目前持有價值'], row['當期回撤']]
        aligns = [center, center, right, right, center, center, right, right, right]
        for col_idx, (val, aln) in enumerate(zip(values, aligns), start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=val)
            cell.font = data_font; cell.alignment = aln; cell.border = thin_border
            if fill: cell.fill = fill
            if col_idx == 9: cell.number_format = '0.00%'

    n             = len(result_df)
    summary_start = n + 3
    total_profit  = round(final_value - total_invested, 2)

    summary_rows = [
        ('績效摘要',         None),
        ('總投入月數（月）', n),
        ('累積總股數（股）', n),
        ('總投入金額',       round(total_invested, 2)),
        ('最終持有市值',     round(final_value, 2)),
        ('總損益',           total_profit),
        ('累積報酬率',       cumulative_return),
        ('最大回撤 MDD',     max_drawdown_1),
        ('年化IRR',          annual_irr),
        ('Sortino Ratio',    sortino_ratio),
    ]

    for i, (label, value) in enumerate(summary_rows):
        r = summary_start + i
        label_cell = ws.cell(row=r, column=1, value=label)
        label_cell.font = summary_font; label_cell.fill = summary_fill
        label_cell.alignment = center; label_cell.border = thin_border
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        val_cell = ws.cell(row=r, column=9, value=value)
        val_cell.font = summary_font; val_cell.fill = summary_fill
        val_cell.border = thin_border; val_cell.alignment = right
        if label in ('年化IRR', '累積報酬率', '最大回撤 MDD'):
            val_cell.number_format = '0.00%'
        elif label == 'Sortino Ratio':
            val_cell.number_format = '0.00'
        elif label not in ('績效摘要', '總投入月數（月）', '累積總股數（股）'):
            val_cell.number_format = '#,##0.00'

    for col_idx in range(1, 10):
        ws.cell(row=summary_start, column=col_idx).font = Font(
            name='Arial', bold=True, size=11, color='FFFFFF')
        ws.cell(row=summary_start, column=col_idx).fill = PatternFill(
            'solid', start_color='2E75B6')

    ws.freeze_panes = 'A2'
    out_filename = f"{ticker.replace('.','_')}_DCA績效.xlsx"
    out_path     = os.path.join(OUTPUT_DIR, out_filename)
    wb.save(out_path)
    print(f"  ✓ {ticker:12} → {out_filename} 已儲存（Sortino={sortino_ratio:.2f}）")
    return out_path

# ====================== Walk-Forward 回測 ======================
def walk_forward_dca(df, ticker):
    results = []
    max_year = df.index.year.max()

    for i, test_start_year in enumerate(TEST_START_YEARS, start=1):
        if test_start_year > max_year:
            break
        test_df = df[df.index.year >= test_start_year]
        if len(test_df) < 2:
            continue

        test_end_year = test_df.index.year.max()
        test_period   = f"{test_start_year}–{test_end_year}"
        window_label  = f"Window_{i:02d}"

        try:
            (_, annual_irr, cumulative_return, _,
             _, max_drawdown, sortino_ratio) = compute_dca(test_df)
            results.append({
                'Window':        window_label,
                'Test_Period':   test_period,
                '累積報酬率':    cumulative_return,
                '年化IRR':       annual_irr,
                '最大回撤 MDD':  max_drawdown,
                'Sortino Ratio': sortino_ratio,
            })
        except Exception as e:
            print(f"  ⚠ {ticker} {window_label} 計算失敗：{e}")

    return results

# ====================== 輸出 Walk-Forward Excel（每個 ticker 獨立一個檔）======================
def save_walkforward_excel(ticker, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "WalkForward"

    header_font  = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    header_fill  = PatternFill('solid', start_color='1F4E79')
    data_font    = Font(name='Arial', size=10)
    alt_fill     = PatternFill('solid', start_color='D6E4F0')
    center       = Alignment(horizontal='center', vertical='center')
    right        = Alignment(horizontal='right',  vertical='center')
    thin_border  = Border(
        left=Side(style='thin',   color='B0C4DE'),
        right=Side(style='thin',  color='B0C4DE'),
        top=Side(style='thin',    color='B0C4DE'),
        bottom=Side(style='thin', color='B0C4DE'),
    )

    col_headers = ['Window', 'Test_Period', '累積報酬率', '年化IRR', '最大回撤 MDD', 'Sortino Ratio']
    col_widths  = [14, 16, 14, 12, 16, 16]
    pct_cols    = {3, 4, 5}

    for col_idx, (h, w) in enumerate(zip(col_headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = center; cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[1].height = 22

    for row_idx, record in enumerate(rows, start=2):
        fill = alt_fill if row_idx % 2 == 0 else None
        values = [record['Window'], record['Test_Period'], record['累積報酬率'],
                  record['年化IRR'], record['最大回撤 MDD'], record['Sortino Ratio']]
        aligns = [center, center, right, right, right, right]
        for col_idx, (val, aln) in enumerate(zip(values, aligns), start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font; cell.alignment = aln; cell.border = thin_border
            if fill: cell.fill = fill
            if col_idx in pct_cols:
                cell.number_format = '0.00%'
            elif col_idx == 6:
                cell.number_format = '0.00'

    ws.freeze_panes = 'A2'

    out_filename = f"{ticker.replace('.','_')}_DCA_WalkForward.xlsx"
    out_path     = os.path.join(OUTPUT_DIR, out_filename)
    wb.save(out_path)
    print(f"  ✓ Walk-Forward → {out_filename} 已儲存")
    return out_path

# ====================== 主程式 ======================
if __name__ == "__main__":
    # 建立輸出資料夾
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    print("=" * 60)
    print("  DCA 定股投資績效分析（每月第一個交易日買 1 股）")
    print(f"  輸出資料夾：{OUTPUT_DIR}")
    print("=" * 60 + "\n")

    for ticker in TICKERS:
        print(f"[ {ticker} ]")
        df = load_csv(ticker)
        if df is None:
            continue

        # 全期逐月明細 Excel
        (result, annual_irr, cumulative_return, total_invested,
         final_value, max_drawdown_1, sortino_ratio) = compute_dca(df)
        save_excel(result, ticker, annual_irr, cumulative_return, total_invested,
                   final_value, max_drawdown_1, sortino_ratio)

        # Walk-Forward
        wf_rows = walk_forward_dca(df, ticker)
        save_walkforward_excel(ticker, wf_rows)
        print(f"  → Walk-Forward 共 {len(wf_rows)} 個 Window\n")

    print("\n完成！請至桌面的 dcaproject 資料夾查看：")
    print("  • 各股票的 *_DCA績效.xlsx（逐月明細）")
    print("  • 各股票的 *_DCA_WalkForward.xlsx（時間序列回測）")