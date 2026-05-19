import pandas as pd
import numpy as np
import numpy_financial as npf
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ====================== 設定 ======================
DESKTOP_PATH = os.path.expanduser("~/Desktop")
OUTPUT_DIR   = os.path.join(DESKTOP_PATH, "bahproject")   # 輸出資料夾
TICKERS      = ['2330.TW', '2454.TW', 'AAPL', 'PEP', 'KO']
RISK_FREE_RATE = 0.017

INIT_USD   = 10_000
INIT_TWD   = 320_000
TW_TICKERS = {'2330.TW', '2454.TW'}

TEST_START_YEARS = list(range(2007, 2027))

# ====================== 讀取 CSV ======================
def load_csv(ticker):
    filename = f"{ticker.replace('.','_')}_20y.csv"
    file_path = os.path.join(DESKTOP_PATH, filename)
    if not os.path.exists(file_path):
        print(f"✗ 找不到檔案：{filename}，請先執行下載程式。")
        return None
    df = pd.read_csv(file_path, parse_dates=['date'], index_col='date')
    df.dropna(subset=['close'], inplace=True)
    return df

# ====================== Buy and Hold 核心計算 ======================
def compute_bah(df, ticker):
    df = df.copy()
    df.index = pd.to_datetime(df.index)

    init_capital = INIT_TWD if ticker in TW_TICKERS else INIT_USD
    buy_date     = df.index[0]
    buy_price    = df.loc[buy_date, 'close']
    shares       = round(init_capital / buy_price, 2)
    actual_cost  = init_capital

    sell_date   = df.index[-1]
    sell_price  = df.loc[sell_date, 'close']
    final_value = shares * sell_price

    total_profit  = final_value - actual_cost
    simple_return = total_profit / actual_cost

    years = (sell_date - buy_date).days / 365.25
    cagr  = (final_value / actual_cost) ** (1 / years) - 1

    total_months = round(years * 12)
    cash_flows   = [0.0] * (total_months + 1)
    cash_flows[0]  = -actual_cost
    cash_flows[-1] = final_value
    monthly_irr  = npf.irr(cash_flows)
    annual_irr   = (1 + monthly_irr) ** 12 - 1

    # 逐月明細
    df['year_month'] = df.index.to_period('M')
    first_trading_days = df.groupby('year_month').apply(
        lambda x: x.index.min(), include_groups=False
    )
    records      = []
    peak_value   = 0.0
    max_drawdown = 0.0

    for month_num, (period, date) in enumerate(first_trading_days.items(), start=1):
        price           = df.loc[date, 'close']
        portfolio_value = round(price * shares, 2)

        if portfolio_value > peak_value:
            peak_value = portfolio_value
        current_drawdown = (portfolio_value - peak_value) / peak_value
        if current_drawdown < max_drawdown:
            max_drawdown = current_drawdown

        records.append({
            '月份':     month_num,
            '日期':     date.strftime('%Y/%m/%d'),
            '當月股價': round(price, 2),
            '持有股數': shares,
            '持有市值': portfolio_value,
            '當期回撤': current_drawdown,
        })

    result_df = pd.DataFrame(records)

    # Sortino Ratio
    prices          = [r['當月股價'] for r in records]
    monthly_returns = [(prices[i] - prices[i-1]) / prices[i-1] for i in range(1, len(prices))]
    monthly_rf      = (1 + RISK_FREE_RATE) ** (1 / 12) - 1
    downside_diffs  = [min(r - monthly_rf, 0) for r in monthly_returns]
    downside_variance  = np.mean([d ** 2 for d in downside_diffs])
    downside_deviation = np.sqrt(downside_variance) * np.sqrt(12)
    sortino_ratio = (annual_irr - RISK_FREE_RATE) / downside_deviation if downside_deviation != 0 else float('nan')

    summary = {
        'buy_date':          buy_date.strftime('%Y/%m/%d'),
        'buy_price':         round(buy_price, 2),
        'sell_date':         sell_date.strftime('%Y/%m/%d'),
        'sell_price':        round(sell_price, 2),
        'shares':            shares,
        'init_capital':      round(init_capital, 2),
        'actual_cost':       round(actual_cost, 2),
        'final_value':       round(final_value, 2),
        'total_profit':      round(total_profit, 2),
        'cumulative_return': simple_return,
        'max_drawdown':      max_drawdown,
        'cagr':              cagr,
        'annual_irr':        annual_irr,
        'sortino_ratio':     sortino_ratio,
        'total_months':      total_months,
        'years':             round(years, 1),
        'currency':          'TWD' if ticker in TW_TICKERS else 'USD',
    }

    return result_df, summary

# ====================== 輸出逐月明細 Excel ======================
def save_excel(result_df, summary, ticker):
    wb = Workbook()
    ws = wb.active
    ws.title = "BuyAndHold績效"

    header_font  = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    header_fill  = PatternFill('solid', start_color='1D5C2E')
    data_font    = Font(name='Arial', size=10)
    alt_fill     = PatternFill('solid', start_color='D6EAD9')
    summary_font = Font(name='Arial', bold=True, size=10, color='1D5C2E')
    summary_fill = PatternFill('solid', start_color='EBF5EC')
    center       = Alignment(horizontal='center', vertical='center')
    right        = Alignment(horizontal='right',  vertical='center')
    thin_border  = Border(
        left=Side(style='thin',   color='A8C5AB'),
        right=Side(style='thin',  color='A8C5AB'),
        top=Side(style='thin',    color='A8C5AB'),
        bottom=Side(style='thin', color='A8C5AB'),
    )
    currency = summary['currency']

    # 頂部買入資訊
    info_rows = [
        ('買入日期',     summary['buy_date']),
        ('買入股價',     summary['buy_price']),
        ('買入股數',     summary['shares']),
        ('初始投入金額', f"{summary['actual_cost']:,.2f} {currency}"),
    ]
    title_font = Font(name='Arial', bold=True, size=12, color='FFFFFF')
    title_fill = PatternFill('solid', start_color='145226')
    for col in range(1, 7):
        c = ws.cell(row=1, column=col)
        c.fill = title_fill; c.border = thin_border
    ws.merge_cells('A1:F1')
    ws['A1'] = f"Buy & Hold 策略 — {ticker}  （初始投入 {summary['actual_cost']:,.0f} {currency}）"
    ws['A1'].font = title_font; ws['A1'].alignment = center
    ws.row_dimensions[1].height = 24

    label_font = Font(name='Arial', bold=True, size=10, color='145226')
    label_fill = PatternFill('solid', start_color='C6E3C9')
    for i, (label, val) in enumerate(info_rows):
        r = i + 2
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = label_font; lc.fill = label_fill
        lc.alignment = center; lc.border = thin_border
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        vc = ws.cell(row=r, column=3, value=val)
        vc.font = data_font; vc.alignment = right; vc.border = thin_border
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)

    # 逐月明細標題
    table_start = len(info_rows) + 3
    headers     = ['月份', '日期', '當月股價', '持有股數', '持有市值', '當期回撤']
    col_widths  = [8, 14, 14, 12, 16, 12]
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=table_start, column=col_idx, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = center; cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[table_start].height = 22

    # 逐月資料
    for row_idx, row in result_df.iterrows():
        excel_row = table_start + 1 + row_idx
        fill = alt_fill if row_idx % 2 == 0 else None
        values = [row['月份'], row['日期'], row['當月股價'],
                  row['持有股數'], row['持有市值'], row['當期回撤']]
        aligns = [center, center, right, center, right, right]
        for col_idx, (val, aln) in enumerate(zip(values, aligns), start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=val)
            cell.font = data_font; cell.alignment = aln; cell.border = thin_border
            if fill: cell.fill = fill
            if col_idx == 6: cell.number_format = '0.00%'

    # 績效摘要
    n             = len(result_df)
    summary_start = table_start + n + 2
    summary_rows  = [
        ('績效摘要',       None),
        ('持有月數',       summary['total_months']),
        ('買入股數',       summary['shares']),
        ('實際投入金額',   summary['actual_cost']),
        ('最終持有市值',   summary['final_value']),
        ('總損益',         summary['total_profit']),
        ('累積報酬率',     summary['cumulative_return']),
        ('最大回撤 MDD',   summary['max_drawdown']),
        ('年化IRR',        summary['annual_irr']),
        ('Sortino Ratio',  summary['sortino_ratio']),
    ]

    for i, (label, value) in enumerate(summary_rows):
        r = summary_start + i
        label_cell = ws.cell(row=r, column=1, value=label)
        label_cell.font = summary_font; label_cell.fill = summary_fill
        label_cell.alignment = center; label_cell.border = thin_border
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        val_cell = ws.cell(row=r, column=6, value=value)
        val_cell.font = summary_font; val_cell.fill = summary_fill
        val_cell.border = thin_border; val_cell.alignment = right
        if label in ('年化IRR', '累積報酬率', '最大回撤 MDD'):
            val_cell.number_format = '0.00%'
        elif label == 'Sortino Ratio':
            val_cell.number_format = '0.00'
        elif label not in ('績效摘要',):
            val_cell.number_format = '#,##0.00'

    for col_idx in range(1, 7):
        ws.cell(row=summary_start, column=col_idx).font = Font(
            name='Arial', bold=True, size=11, color='FFFFFF')
        ws.cell(row=summary_start, column=col_idx).fill = PatternFill(
            'solid', start_color='2E7D32')

    ws.freeze_panes = f'A{table_start + 1}'
    out_filename = f"{ticker.replace('.','_')}_BuyAndHold績效.xlsx"
    out_path     = os.path.join(OUTPUT_DIR, out_filename)
    wb.save(out_path)
    print(f"  ✓ {ticker:12} → {out_filename}  IRR={summary['annual_irr']:.2%}  Sortino={summary['sortino_ratio']:.2f}")
    return out_path

# ====================== Walk-Forward 回測 ======================
def walk_forward_bah(df, ticker):
    results  = []
    max_year = df.index.year.max()

    for i, test_start_year in enumerate(TEST_START_YEARS, start=1):
        if test_start_year > max_year:
            break
        test_df = df[df.index.year >= test_start_year]
        if len(test_df) < 2:
            continue

        test_end_year = test_df.index.year.max()
        test_period   = f"{test_start_year}–{test_end_year}"
        window_label  = f"Window_{i:02d}"

        try:
            _, summary = compute_bah(test_df, ticker)
            results.append({
                'Window':        window_label,
                'Test_Period':   test_period,
                '累積報酬率':    summary['cumulative_return'],
                '年化IRR':       summary['annual_irr'],
                '最大回撤 MDD':  summary['max_drawdown'],
                'Sortino Ratio': summary['sortino_ratio'],
            })
        except Exception as e:
            print(f"  ⚠ {ticker} {window_label} 計算失敗：{e}")

    return results

# ====================== 輸出 Walk-Forward Excel ======================
def save_walkforward_excel(ticker, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "WalkForward"

    header_font  = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    header_fill  = PatternFill('solid', start_color='1D5C2E')
    data_font    = Font(name='Arial', size=10)
    alt_fill     = PatternFill('solid', start_color='D6EAD9')
    center       = Alignment(horizontal='center', vertical='center')
    right        = Alignment(horizontal='right',  vertical='center')
    thin_border  = Border(
        left=Side(style='thin',   color='A8C5AB'),
        right=Side(style='thin',  color='A8C5AB'),
        top=Side(style='thin',    color='A8C5AB'),
        bottom=Side(style='thin', color='A8C5AB'),
    )

    col_headers = ['Window', 'Test_Period', '累積報酬率', '年化IRR', '最大回撤 MDD', 'Sortino Ratio']
    col_widths  = [14, 16, 14, 12, 16, 16]
    pct_cols    = {3, 4, 5}

    for col_idx, (h, w) in enumerate(zip(col_headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = center; cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[1].height = 22

    for row_idx, record in enumerate(rows, start=2):
        fill = alt_fill if row_idx % 2 == 0 else None
        values = [record['Window'], record['Test_Period'], record['累積報酬率'],
                  record['年化IRR'], record['最大回撤 MDD'], record['Sortino Ratio']]
        aligns = [center, center, right, right, right, right]
        for col_idx, (val, aln) in enumerate(zip(values, aligns), start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font; cell.alignment = aln; cell.border = thin_border
            if fill: cell.fill = fill
            if col_idx in pct_cols:
                cell.number_format = '0.00%'
            elif col_idx == 6:
                cell.number_format = '0.00'

    ws.freeze_panes = 'A2'
    out_filename = f"{ticker.replace('.','_')}_BuyAndHold_WalkForward.xlsx"
    out_path     = os.path.join(OUTPUT_DIR, out_filename)
    wb.save(out_path)
    print(f"  ✓ Walk-Forward → {out_filename} 已儲存")
    return out_path

# ====================== 主程式 ======================
if __name__ == "__main__":
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    print("=" * 60)
    print("  Buy & Hold 績效分析")
    print(f"  美股初始資金：${INIT_USD:,}  |  台股初始資金：NT${INIT_TWD:,}")
    print(f"  輸出資料夾：{OUTPUT_DIR}")
    print("=" * 60 + "\n")

    for ticker in TICKERS:
        print(f"[ {ticker} ]")
        df = load_csv(ticker)
        if df is None:
            continue

        # 全期逐月明細 Excel
        result_df, summary = compute_bah(df, ticker)
        save_excel(result_df, summary, ticker)

        # Walk-Forward
        wf_rows = walk_forward_bah(df, ticker)
        save_walkforward_excel(ticker, wf_rows)
        print(f"  → Walk-Forward 共 {len(wf_rows)} 個 Window\n")

    print("完成！請至桌面的 bahproject 資料夾查看：")
    print("  • 各股票的 *_BuyAndHold績效.xlsx（逐月明細）")
    print("  • 各股票的 *_BuyAndHold_WalkForward.xlsx（時間序列回測）")